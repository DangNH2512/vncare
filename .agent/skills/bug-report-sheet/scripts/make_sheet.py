#!/usr/bin/env python3
"""Build the QC bug tracker workbook for Da Nang Connect.

Usage:
    python make_sheet.py --rows tracker.csv --out out.xlsx
                         [--sheet bug-tracker-DNC] [--blank-rows 60]

One workbook, two sheets: the tracker itself plus a "Tom tat" dashboard that
counts itself with COUNTIFS. The layout is the QC house format of this project:
20 columns, a Blocker/Major/Minor scale, a fixed Dev Status vocabulary and
dropdowns that reach far below the last data row so newly typed rows already
have them.

Product-specific notes:
    * Platform covers Web (apps/web), Android and iOS (apps/mobile) plus "All".
      A defect that reproduces on several platforms is written "Web, iOS" and is
      counted once per platform, so platform totals may exceed the row count.
    * The UI ships in English first and Vietnamese second, so a locale-specific
      defect must say which locale it was seen in, inside the "Environment"
      column (e.g. "Web Chrome 129 - locale VI").
"""

from __future__ import annotations

import argparse
import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── House format: 18 columns, A..R (column A is the running number) ──────────
COLUMNS = [
    # "Xong chưa?" answers the one question the owner opens this file for — is it
    # done — because eight Dev Status values turn that into a lookup instead of a
    # glance. Dev Status stays for the detail.
    " ", "Xong chưa?",
    # "Loại" separates a real defect from a feature request: tester batches for a
    # community app are always a mix of "RSVP nhân đôi" (bug) and "cho phép lọc
    # theo khung giờ" (đề xuất). Reading "Xong" next to a request for a
    # not-yet-built feature reads as "the bug is fixed" — this column says up
    # front what kind of row it is.
    "Loại", "Revision/Build", "Platform", "Thời điểm test", "Màn hình", "Content",
    "Pre-condition", "Expected", "Actual", "Environment", "Severity",
    "Proof (ảnh tester)", "Dev Status", "Fix commit", "Proof-of-fix (ảnh sau fix)",
    "Verify", "Phản hồi Dev (chi tiết)", "Phản hồi QC (chi tiết)",
]
WIDTHS = [6, 13, 15, 19.14, 14.29, 17.29, 26, 56.86, 22.71, 40.57, 30, 21, 12, 38, 12, 14,
          22.29, 10, 73.14, 34.29]

# Vocabularies — the only accepted values; the summary sheet counts these exact
# strings, so changing one here means changing it in write_summary() too.
LOAI = ["Bug", "Đề xuất", "Không phải bug", "Cảnh báo"]
PLATFORMS = ["All", "Web", "Android", "iOS"]
SEVERITIES = ["Blocker", "Major", "Minor"]
DEV_STATUS = ["New", "Fixed", "Wont-fix", "Cannot-repro", "Đã trả lời",
              "In progress", "Cần PO xác nhận", "Chấp nhận tạm - làm sau"]
VERIFY = ["OK", "Fail"]

HEADER_FILL = PatternFill("solid", fgColor="2E7D32")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
SEV_FILL = {"Blocker": PatternFill("solid", fgColor="C00000"),
            "Major": PatternFill("solid", fgColor="F4B183"),
            "Minor": PatternFill("solid", fgColor="FFE699")}
SEV_FONT = {"Blocker": Font(color="FFFFFF", bold=True),
            "Major": Font(color="7F2704", bold=True),
            "Minor": Font(color="7F6000", bold=True)}
DONE_FILL = {
    "Xong": PatternFill("solid", fgColor="C6E0B4"),
    "Chưa xong": PatternFill("solid", fgColor="FFC7CE"),
}
DONE_FONT = {
    "Xong": Font(color="375623", bold=True),
    "Chưa xong": Font(color="9C0006", bold=True),
}
# Blue for an actual defect, purple for a feature request, grey for the other
# two — neither is "wrong", they just aren't something to fix in code.
LOAI_FILL = {
    "Bug": PatternFill("solid", fgColor="9DC3E6"),
    "Đề xuất": PatternFill("solid", fgColor="D9B3F5"),
    "Không phải bug": PatternFill("solid", fgColor="D9D9D9"),
    "Cảnh báo": PatternFill("solid", fgColor="FFD966"),
}
LOAI_FONT = {
    "Bug": Font(color="1F3864", bold=True),
    "Đề xuất": Font(color="4C2A66", bold=True),
    "Không phải bug": Font(color="404040", bold=True),
    "Cảnh báo": Font(color="7F5F00", bold=True),
}

STATUS_FILL = {"New": PatternFill("solid", fgColor="FCE4D6"),
               "In progress": PatternFill("solid", fgColor="DDEBF7"),
               "Fixed": PatternFill("solid", fgColor="C6E0B4"),
               "Đã trả lời": PatternFill("solid", fgColor="FFF2CC"),
               "Wont-fix": PatternFill("solid", fgColor="EDEDED"),
               "Cannot-repro": PatternFill("solid", fgColor="EDEDED"),
               "Cần PO xác nhận": PatternFill("solid", fgColor="E4DFEC"),
               "Chấp nhận tạm - làm sau": PatternFill("solid", fgColor="EDEDED")}
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
VALIDATION_LIMIT = 1000  # dropdowns reach far below the data so new rows keep them


def read_rows(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8-sig", newline="") as fh:
        return [dict(r) for r in csv.DictReader(fh)]


def write_tracker(wb: Workbook, sheet: str, rows: list[dict[str, str]],
                  blank_rows: int) -> int:
    ws = wb.create_sheet(sheet)
    ws.append(COLUMNS)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    ws.row_dimensions[1].height = 32

    for idx, row in enumerate(rows, start=1):
        ws.append([float(idx) if col == " " else row.get(col, "") for col in COLUMNS])

    last_data = ws.max_row
    last_row = last_data + blank_rows

    for i, width in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    centered = {"Platform", "Severity", "Dev Status", "Verify", " ", "Xong chưa?", "Loại"}
    def needed_height(r: int) -> float:
        """Tallest cell in the row decides the row height."""
        lines = 1
        for c, col in enumerate(COLUMNS, start=1):
            v = ws.cell(row=r, column=c).value
            if not isinstance(v, str) or not v:
                continue
            width = WIDTHS[c - 1] if c - 1 < len(WIDTHS) else 26
            n = 0
            for seg in v.split("\n"):
                n += max(1, -(-len(seg) // max(8, int(width * 0.95))))
            lines = max(lines, n)
        return min(430, max(48, lines * 13.2))

    for r in range(2, last_row + 1):
        ws.row_dimensions[r].height = needed_height(r) if r <= last_data else 30
        for c, col in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = Alignment(
                vertical="center" if col in centered else "top",
                horizontal="center" if col in centered else "general",
                wrap_text=True)
        sev = ws.cell(row=r, column=COLUMNS.index("Severity") + 1)
        if sev.value in SEV_FILL:
            sev.fill = SEV_FILL[sev.value]
            sev.font = SEV_FONT[sev.value]
        st = ws.cell(row=r, column=COLUMNS.index("Dev Status") + 1)
        if st.value in STATUS_FILL:
            st.fill = STATUS_FILL[st.value]
        done = ws.cell(row=r, column=COLUMNS.index("Xong chưa?") + 1)
        if done.value in DONE_FILL:
            done.fill = DONE_FILL[done.value]
            done.font = DONE_FONT[done.value]
        loai = ws.cell(row=r, column=COLUMNS.index("Loại") + 1)
        if loai.value in LOAI_FILL:
            loai.fill = LOAI_FILL[loai.value]
            loai.font = LOAI_FONT[loai.value]

    for col, values in (("Loại", LOAI), ("Platform", PLATFORMS), ("Severity", SEVERITIES),
                        ("Dev Status", DEV_STATUS), ("Verify", VERIFY)):
        letter = get_column_letter(COLUMNS.index(col) + 1)
        dv = DataValidation(type="list", formula1='"' + ",".join(values) + '"',
                            allow_blank=True, errorStyle="warning",
                            errorTitle="Ngoài danh mục",
                            error="Giá trị không có trong danh mục chuẩn. Vẫn nhập được "
                                  "(vd Platform 'Web, iOS' khi lỗi gặp ở cả hai).")
        ws.add_data_validation(dv)
        dv.add(f"{letter}2:{letter}{VALIDATION_LIMIT}")

    # Freeze through "Loại" (not a hardcoded "C2"): STT / Xong / Loại stay pinned
    # while scrolling right, same intent as before, just one column further now.
    freeze_col = get_column_letter(COLUMNS.index("Loại") + 2)
    ws.freeze_panes = f"{freeze_col}2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{last_row}"
    return last_row


def _letter(name: str) -> str:
    """Column letter for a header — hardcoding these broke when a column moved."""
    return get_column_letter(COLUMNS.index(name) + 1)


def write_summary(wb: Workbook, sheet: str, last_row: int) -> None:
    ws = wb.create_sheet("Tóm tắt", 0)
    q = f"'{sheet}'"
    lim = VALIDATION_LIMIT

    PLATFORM, SEVERITY, DEV_STATUS, VERIFY = (
        _letter("Platform"), _letter("Severity"), _letter("Dev Status"), _letter("Verify"))
    DONE = _letter("Xong chưa?")
    LOAI_COL = _letter("Loại")
    ws.column_dimensions["A"].width = 30
    for letter in "BCDEFGHIJK":
        ws.column_dimensions[letter].width = 15

    ws["A1"] = f"BẢNG TÓM TẮT BUG — tự đếm từ sheet '{sheet}'"
    ws["A1"].font = Font(bold=True, size=14, color="1B5E20")
    ws["A2"] = ("⚠️ Lỗi gặp ở nhiều nền tảng (vd 'Web, iOS') được đếm ở MỖI nền tảng "
                "liên quan → tổng theo platform có thể > số dòng bug.")
    ws["A2"].font = Font(italic=True, color="7F7F7F")

    ws["A4"] = "Tổng số bug (số dòng):"
    ws["A4"].font = Font(bold=True)
    ws["B4"] = f"=COUNTA({q}!A2:A{lim})"
    ws["B4"].font = Font(bold=True, size=12)
    # Headline first: XONG / CHƯA XONG is what gets read, everything below is detail.
    ws["A5"] = "XONG:"
    ws["B5"] = f'=COUNTIF({q}!${DONE}$2:${DONE}${lim},"Xong")'
    ws["A6"] = "CHƯA XONG:"
    ws["B6"] = f'=COUNTIF({q}!${DONE}$2:${DONE}${lim},"Chưa xong")'
    for ref in ("A5", "B5", "A6", "B6"):
        ws[ref].font = Font(bold=True, size=12)
    ws["B5"].fill = PatternFill("solid", fgColor="C6E0B4")
    ws["B6"].fill = PatternFill("solid", fgColor="FFC7CE")

    def table(top: int, title: str, head: list[str], rows: list[tuple[str, list[str]]]) -> int:
        ws.cell(row=top, column=1, value=title).font = Font(bold=True, size=12, color="FFFFFF")
        for c in range(1, len(head) + 1):
            ws.cell(row=top, column=c).fill = HEADER_FILL
        r = top + 1
        for c, name in enumerate(head, start=1):
            cell = ws.cell(row=r, column=c, value=name)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="E2EFDA")
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        for label, formulas in rows:
            r += 1
            cell = ws.cell(row=r, column=1, value=label)
            cell.border = BORDER
            cell.font = Font(bold=True)
            for c, formula in enumerate(formulas, start=2):
                fc = ws.cell(row=r, column=c, value=formula)
                fc.border = BORDER
                fc.alignment = Alignment(horizontal="center")
        return r + 2

    # "Phân loại" answers "how many of these rows are actually defects?".
    # Feature requests, false alarms and doc-accuracy risks are not defects, and
    # lumping them into "Xong/Chưa xong" makes that invisible.
    rows_loai = [(l, [f'=COUNTIF({q}!${LOAI_COL}$2:${LOAI_COL}${lim},"{l}")']) for l in LOAI]
    nxt = table(8, "PHÂN LOẠI — bao nhiêu dòng thực sự là bug", ["Loại", "Số dòng"], rows_loai)

    ws.cell(row=nxt, column=1, value="Đã Fixed:").font = Font(bold=True)
    ws.cell(row=nxt, column=2,
            value=f'=COUNTIF({q}!${DEV_STATUS}$2:${DEV_STATUS}${lim},"Fixed")').font = Font(bold=True, size=12)
    ws.cell(row=nxt + 1, column=1, value="Đã QC verify OK:").font = Font(bold=True)
    ws.cell(row=nxt + 1, column=2,
            value=f'=COUNTIF({q}!${VERIFY}$2:${VERIFY}${lim},"OK")').font = Font(bold=True, size=12)
    ws.cell(row=nxt + 2, column=1, value="Tỉ lệ đã đóng (verify OK):").font = Font(bold=True)
    rate_cell = ws.cell(row=nxt + 2, column=2,
                         value=(f'=IF(COUNTA({q}!A2:A{lim})=0,0,'
                                f'COUNTIF({q}!${VERIFY}$2:${VERIFY}${lim},"OK")/COUNTA({q}!A2:A{lim}))'))
    rate_cell.font = Font(bold=True, size=12)
    rate_cell.number_format = "0%"
    nxt += 4

    # BẢNG 1 — Platform × Dev Status. Đếm ĐỦ mọi trạng thái trong DEV_STATUS: bỏ sót
    # một giá trị nào là các dòng mang trạng thái đó biến mất khỏi tổng.
    statuses = list(DEV_STATUS)
    rows1 = []
    for plat in ["Web", "Android", "iOS", "All"]:
        formulas = [f'=COUNTIFS({q}!${PLATFORM}$2:${PLATFORM}${lim},"*{plat}*",{q}!${DEV_STATUS}$2:${DEV_STATUS}${lim},"{s}")'
                    for s in statuses]
        formulas.append(f'=COUNTIF({q}!${PLATFORM}$2:${PLATFORM}${lim},"*{plat}*")')
        rows1.append((plat, formulas))
    nxt = table(nxt, "BẢNG 1 — Platform × Trạng thái (Dev Status)",
                ["Platform"] + statuses + ["Tổng"], rows1)

    # BẢNG 2 — Severity × Platform
    rows2 = []
    for sev in SEVERITIES:
        formulas = [f'=COUNTIFS({q}!${SEVERITY}$2:${SEVERITY}${lim},"{sev}",{q}!${PLATFORM}$2:${PLATFORM}${lim},"*{p}*")'
                    for p in ["Web", "Android", "iOS", "All"]]
        formulas.append(f'=COUNTIF({q}!${SEVERITY}$2:${SEVERITY}${lim},"{sev}")')
        rows2.append((sev, formulas))
    nxt = table(nxt, "BẢNG 2 — Severity × Platform",
                ["Severity", "Web", "Android", "iOS", "All", "Tổng"], rows2)

    # BẢNG 3 — Severity × Dev Status (còn bao nhiêu Blocker/Major chưa fix)
    rows3 = []
    for sev in SEVERITIES:
        formulas = [f'=COUNTIFS({q}!${SEVERITY}$2:${SEVERITY}${lim},"{sev}",{q}!${DEV_STATUS}$2:${DEV_STATUS}${lim},"{s}")'
                    for s in statuses]
        formulas.append(f'=COUNTIF({q}!${SEVERITY}$2:${SEVERITY}${lim},"{sev}")')
        rows3.append((sev, formulas))
    table(nxt, "BẢNG 3 — Severity × Trạng thái (còn bao nhiêu Blocker/Major chưa fix)",
          ["Severity"] + statuses + ["Tổng"], rows3)

    ws.freeze_panes = "A3"


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--rows", required=True, type=Path)
    ap.add_argument("--out", required=True, type=Path)
    ap.add_argument("--sheet", default="bug-tracker-DNC")
    ap.add_argument("--blank-rows", type=int, default=60)
    args = ap.parse_args()

    rows = read_rows(args.rows)
    wb = Workbook()
    wb.remove(wb.active)
    last = write_tracker(wb, args.sheet, rows, args.blank_rows)
    write_summary(wb, args.sheet, last)

    args.out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.out)
    print(f"wrote {args.out} — {len(rows)} bug + {args.blank_rows} dòng trống "
          f"(sheet '{args.sheet}' + 'Tóm tắt')")


if __name__ == "__main__":
    main()
